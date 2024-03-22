# from typing_extensions import dataclass_transform
from selenium.common import TimeoutException
import time
import random
from classspider import MySpider
from multiprocessing import Process
from pickle import TRUE
import re
from selenium import webdriver
import time
from selenium.webdriver import Chrome
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
from selenium.webdriver.support.wait import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC


import pandas as pd
import openpyxl
from tqdm import tqdm
import json
import pandas as pd
import numpy as np


# 所有的有数据的分类号

myspider = MySpider()
def contains_chinese(text):
    pattern = re.compile(r'[\u4e00-\u9fa5]')  # 匹配中文字符的正则表达式
    match = re.search(pattern, text)  # 在文本中搜索匹配的内容
    return match is not None
# 爬取数据
def spider(labelnum,paper_num,cc):
    flag = True
    data = openpyxl.Workbook()  # 新建工作簿
    data.create_sheet('Sheet1')  # 添加页
    table = data.active  # 获得当前活跃的工作页，默认为第一个工作页
    chrome_options = webdriver.ChromeOptions()
    chrome_options.add_argument('--headless')
    chrome_options.add_argument('--no-sandbox')
    chrome_options.add_argument('--disable-gpu')
    chrome_options.add_argument('--disable-dev-shm-usage')
    web = webdriver.Chrome(options=chrome_options)
    try:
        url = 'https://s.wanfangdata.com.cn/advanced-search/paper'
        # 暂停随机的等待时间
        time.sleep(5)
        web.get(url)
    except:
        print("打开页面失败")
        return
    try:
        # 暂停随机的等待时间
        time.sleep(5)
        myspider.Search_id(web,cc)
        # 暂停随机的等待时间
        time.sleep(5)
        # choose不成功哇好像 成功了
        myspider.U_Choose(web,cc)
        # 暂停随机的等待时间
        time.sleep(5)
        myspider.u_Class(web,cc)
        # 暂停随机的等待时间
        time.sleep(5)
        # 再单独点击一下呢
        myspider.Search(web,cc)
    except:
        print("*"*20+cc+"*"*20)
        print("死了，下一个")
        return
    c = 0
    while True:
        tag = 0
        print("*"*20+cc+"*"*20)
        try:
            # tr_list = '/html/body/div[5]/div/div[2]/div[2]/div[2]/div[2]/div[2]/div/div'
            # results = web.find_element(By.XPATH,tr_list)
            results = web.find_elements(By.CLASS_NAME, "result-list")
            print("获取到论文列表成功")
        except:
            print("没有获取到论文列表")
            break

        # 每一页的论文列表 /html/body/div[5]/div/div[2]/div[2]/div[2]/div[2]/div[2]/div/div/div[3]/div[1] （最后一个div1-20）
        # 后面跟标题连接的span /div/div[1]/div[2]/span[2]接上
        # 这里循环20次点击论文标题
        for i in range(20):#0-19
            tag = 0
            titlespan = "/html/body/div[5]/div/div[2]/div[2]/div[2]/div[2]/div[2]/div/div/div[3]/div["+str(i+1)+"]/div/div[1]/div[2]/span[2]"
            # 如果没有论文标题也就是没有这个titlespan就直接break掉里外两个循环
            try:
                WebDriverWait(web, 5).until(
                    EC.presence_of_element_located((By.XPATH, titlespan))
                )
                print("定位论文标题成功")
            except:
                flag = False
                print("定位论文标题失败")
                break
            # 如果有就正常点击这个论文标题来到新页面
            try:
                time.sleep(5)
                myspider.Click_title(web,titlespan,cc)
                # 这里打开新页面了记得返回
                if myspider.is_childpage(web) == True:
                    time.sleep(5)
                    web.switch_to.window(web.window_handles[1])
                    print(web.window_handles)
                else:
                    print('未打开新窗口')
                    continue
            except:
                    print("***论文页面打开失败")
                    time.sleep(2)
                    continue
            # 进入新页面,题名\摘要\作者\关键词\文献来源\机构\年\卷\期\页码\issn\中图分类号

            # 摘要如果有打开点一下
            # 题名\摘要\作者\关键词\文献来源\机构\年\卷\期\页码\issn\中图分类号
            wait_time = 10  # 最大等待时间（秒）
            try:
                WebDriverWait(web, 10).until(
                    EC.presence_of_element_located((By.CLASS_NAME, 'detailTitle'))
                )
                web.find_element(By.CLASS_NAME, 'detailTitle')
                print("页面正常显示")
            except:
                print("页面没有正常显示")
                web.close()  # 关闭子网页
                web.switch_to.window(web.window_handles[0])  # 切换到原网页
                continue
            titleclass = 'detailTitle'
            absclass = 'summary'
            autherclass = 'author'
            keywordclass = 'keyword'
            sourceclass = 'periodicalName'
            orgnaclass = 'organization'
            yearclass = 'publishData'
            pageclass = 'pages'
            issnclass = 'periodicalDataItem'
            clsclass = 'classify'
            try:
                title = web.find_element(By.CLASS_NAME,titleclass).text
                
                abstrat = web.find_element(By.CLASS_NAME,absclass).text
                abstrat = abstrat.replace('摘要：','').strip()
                try:
                    source = web.find_element(By.CLASS_NAME,sourceclass).text
                except:
                    source = ''
                # 判断摘要是否存在中文
                if contains_chinese(abstrat)==False:
                    # 如果不存在直接不存了
                    continue
                auther = web.find_element(By.CLASS_NAME,autherclass).text
                names_list = re.findall(r'[^\d\n]+', auther)  # 使用 findall 方法找到所有满足条件的部分

                # 去除空白和换行，并用分号隔开每个名字
                auther = ';'.join(name.strip() for name in names_list if name.strip())

                keyword_container = web.find_element(By.CLASS_NAME,keywordclass)
                anchor_elements = keyword_container.find_elements(By.TAG_NAME,'span')
                keywords_list = []


                # 遍历所有 <a> 元素，将内容添加到列表中
                for anchor_element in anchor_elements:
                    keyword = anchor_element.text.strip()   # 获取 <a> 元素的文本内容，并去除首尾空格
                    keywords_list.append(keyword)  # 将内容添加到列表中

                # 将关键词列表用分号隔开，并保存到 keywords 变量中
                keywords = ';'.join(keywords_list)
                keywords = keywords.replace("关键词：;",'').strip()
                cls_container = web.find_element(By.CLASS_NAME,clsclass)
                cls_anchor_elements = cls_container.find_elements(By.TAG_NAME,'span')
                cls_list = []
                for cls_anchor_element in cls_anchor_elements:
                    cls = cls_anchor_element.text.strip()
                    cls_list.append(cls)
                cls = ';'.join(cls_list)
                cls = cls.replace("机标分类号：;", '').strip()
                print(cls)
                # source = web.find_element(By.XPATH,sour).text
                organ = web.find_element(By.CLASS_NAME,orgnaclass).text
                nianjuanqi = web.find_element(By.CLASS_NAME,yearclass).text
                content_without_prefix = nianjuanqi.replace("年,卷(期)：", "")

                # 使用逗号和括号进行分割
                year, volume_issue = content_without_prefix.split(",")

                # 进一步分割卷和期，并去除括号
                volume, issue = volume_issue.split("(")
                issue = issue.replace(")", "")
                page = web.find_element(By.CLASS_NAME,pageclass).text
                page = page.replace("页数：", '').strip()
                try:
                    issn = web.find_element(By.CLASS_NAME,issnclass).text
                    issn = issn.replace('ISSN：', '')
                except:
                    issn = ''
                    issn = issn.replace('ISSN：', '')



                print(title,abstrat,auther,keywords,source,organ,year,volume,issue,page,issn,cls,cc)

                # 保存下来

                table.cell(paper_num+1,1,paper_num) # type: ignore
                table.cell(paper_num+1,2,title if len(title)!= 0 else " ") # type: ignore
                table.cell(paper_num+1,3,abstrat if len(abstrat)!= 0 else " ") # type: ignore
                table.cell(paper_num+1,4,auther if len(auther)!= 0 else " ") # type: ignore
                table.cell(paper_num+1,5,keywords if len(keywords)!= 0 else " ") # type: ignore
                table.cell(paper_num + 1, 6, source if len(source) != 0 else " ")  # type: ignore
                # table.cell(paper_num+1,6,source if len(source)!= 0 else " ")
                table.cell(paper_num+1,7,organ if len(organ)!= 0 else " ") # type: ignore
                table.cell(paper_num+1,8,year if len(year)!= 0 else " ") # type: ignore
                table.cell(paper_num+1,9,volume if len(volume)!= 0 else " ")
                table.cell(paper_num+1,10,issue if len(issue)!= 0 else " ")
                table.cell(paper_num+1,11,page if len(page)!= 0 else " ") # type: ignore
                table.cell(paper_num+1,12,issn if len(issn)!= 0 else " ") # type: ignore
                table.cell(paper_num+1,13,cls if len(cls)!= 0 else " ") # type: ignore


                paper_num += 1
            except:
                print("信息获取失败,试试只获取几个")

                tag =1
                # title = web.find_element(By.XPATH,titlexpath).text
                # abstrat = web.find_element(By.XPATH,absxpath).text
                # keywords = web.find_element(By.XPATH,keywordxpath).text
                # cls = web.find_element(By.XPATH,clsxpath).text
                # print(title,abstrat,keywords,cls)
                # print(title)
            try:
                data.save('wanfang' + cc + '_' + str(labelnum) + '.xlsx')
            except:
                print("保存失败")
                tag=1

            if tag == 0:
                c += 1
            print(cc + ":目前已有" + str(c) + "个数据")
            if c >= 100:
                return
            try:
                if myspider.is_childpage(web) == True:
                    time.sleep(5)
                    web.close()# 关闭子网页
                    time.sleep(5)
                    web.switch_to.window(web.window_handles[0]) # 切换到原网页
                    # print(web.window_handles)
                    # print("返回列表页")
            except:
                print("返回列表页失败")

        if flag==False:
            flag = True
            break

        # 翻页
        try:
            if myspider.Next_page(web) == True:
                # web.find_element_by_xpath('/html/body/div[5]/div/div[2]/div[4]/div[2]/div[2]/div[2]/div/div/div[5]/span[6]').click()
                # web.find_element(By.XPATH,'/html/body/div[5]/div/div[2]/div[3]/div[2]/div[2]/div[2]/div/div/div[5]/span[9]').send_keys(Keys.RIGHT)
                element = WebDriverWait(web, 10).until(
                    EC.element_to_be_clickable(
                        (By.CLASS_NAME, 'next'))
                )
                time.sleep(5)
                element.click()
                print("下一页")
                time.sleep(5)
            else:
                return
        except:
            print("点击下一页失败")
            return

    web.close()


def ccspider(lablenum,endnum):
    paper_num = 1
    start = int(lablenum)
    end = int(endnum)
    numbercount = int(lablenum)
    for i in range(10):
        uclass = open('aluuu' + str(i) + '.txt', 'r', encoding='utf-8').read().split("\n")
        for  cc in uclass[start:end]:
            print("第"+str(numbercount)+"个标签"+cc,end=" ")
            spider(lablenum,paper_num,cc)
            numbercount+=1


def main():
    p1 = Process(target=ccspider, args=(str(0),str(1),))
    p1.start()

    time.sleep(5)
    p2 = Process(target=ccspider, args=(str(1),str(2),))
    p2.start()
    time.sleep(5)
    p3 = Process(target=ccspider, args=(str(2),str(3),))
    p3.start()
    time.sleep(5)
    p4 = Process(target=ccspider, args=(str(3),str(4),))
    p4.start()
    time.sleep(5)
    p5 = Process(target=ccspider, args=(str(4),str(5),))
    p5.start()
    p6 = Process(target=ccspider, args=(str(5), str(6),))
    p6.start()
    time.sleep(5)
    p7 = Process(target=ccspider, args=(str(6), str(7),))
    p7.start()
    time.sleep(5)
    p8 = Process(target=ccspider, args=(str(7), str(8),))
    p8.start()
    time.sleep(5)
    p9 = Process(target=ccspider, args=(str(8), str(9),))
    p9.start()
    time.sleep(5)
    p10 = Process(target=ccspider, args=(str(9), str(10),))
    p10.start()

    return
if __name__ == '__main__':
    # a = input()
     #ccspider(0,1)
      main()
    # spider(1)

